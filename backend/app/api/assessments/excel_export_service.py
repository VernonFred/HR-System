"""提交记录 Excel 导出服务."""
from typing import Optional

from sqlmodel import Session, select

from app.models_assessment import Questionnaire, Submission


async def export_submissions_to_excel(
    session: Session,
    category: Optional[str] = None,
    questionnaire_id: Optional[int] = None
) -> bytes:
    """导出提交记录为Excel文件."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from sqlmodel import select

    # 构建查询
    query = select(Submission)

    if questionnaire_id:
        query = query.where(Submission.questionnaire_id == questionnaire_id)

    if category:
        q_statement = select(Questionnaire.id).where(Questionnaire.category == category)
        questionnaire_ids = session.exec(q_statement).all()
        if questionnaire_ids:
            query = query.where(Submission.questionnaire_id.in_(questionnaire_ids))

    submissions = session.exec(query).all()

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "提交记录"

    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # 写入表头
    headers = ["序号", "姓名", "电话", "问卷", "得分", "等级", "状态", "提交时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20

    # 写入数据
    for row, sub in enumerate(submissions, 2):
        ws.cell(row=row, column=1, value=row - 1).border = thin_border
        ws.cell(row=row, column=2, value=sub.candidate_name or "").border = thin_border
        ws.cell(row=row, column=3, value=sub.candidate_phone or "").border = thin_border
        ws.cell(row=row, column=4, value=sub.questionnaire_name or "").border = thin_border
        ws.cell(row=row, column=5, value=sub.total_score).border = thin_border
        ws.cell(row=row, column=6, value=sub.grade or "").border = thin_border
        ws.cell(row=row, column=7, value="已完成" if sub.status == "completed" else "进行中").border = thin_border
        ws.cell(row=row, column=8, value=sub.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if sub.submitted_at else "").border = thin_border

    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()
