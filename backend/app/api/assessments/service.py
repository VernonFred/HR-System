"""问卷/测评管理 - 业务逻辑."""
from datetime import datetime
from typing import List, Tuple, Optional
from sqlmodel import Session, select, func, and_
import random
import string

from app.models_assessment import Questionnaire, Assessment, Submission
from app.models import Candidate
from app.professional_scoring import (
    score_professional_assessment,
    score_custom_questionnaire,
    ProfessionalScoringError
)
from app.custom_scoring import calculate_custom_questionnaire_score


# ========== 问卷管理 ==========

async def get_questionnaires(
    session: Session, skip: int = 0, limit: int = 100, category: Optional[str] = None
) -> Tuple[List[Questionnaire], int]:
    """获取问卷列表，支持按category过滤.
    
    Args:
        session: 数据库会话
        skip: 跳过数量
        limit: 返回数量
        category: 问卷分类过滤（professional/scored/survey/custom）
                  'custom' 表示获取所有非professional的问卷（scored + survey）
    """
    # 构建查询条件
    base_query = select(Questionnaire)
    count_query = select(func.count()).select_from(Questionnaire)
    
    if category:
        if category == 'custom':
            # ⭐ custom类别：获取所有非professional的问卷（scored + survey）
            base_query = base_query.where(Questionnaire.category.in_(['scored', 'survey']))
            count_query = count_query.where(Questionnaire.category.in_(['scored', 'survey']))
        else:
            base_query = base_query.where(Questionnaire.category == category)
            count_query = count_query.where(Questionnaire.category == category)
    
    total = session.scalar(count_query)
    statement = base_query.offset(skip).limit(limit).order_by(Questionnaire.created_at.desc())
    questionnaires = session.exec(statement).all()
    return list(questionnaires), total or 0


async def get_questionnaire(session: Session, questionnaire_id: int) -> Optional[Questionnaire]:
    """获取问卷详情."""
    return session.get(Questionnaire, questionnaire_id)


async def create_questionnaire(session: Session, data: dict) -> Questionnaire:
    """创建问卷."""
    questionnaire = Questionnaire(**data)
    session.add(questionnaire)
    session.commit()
    session.refresh(questionnaire)
    return questionnaire


async def update_questionnaire(
    session: Session, questionnaire_id: int, data: dict
) -> Optional[Questionnaire]:
    """更新问卷."""
    questionnaire = session.get(Questionnaire, questionnaire_id)
    if not questionnaire:
        return None
    
    for key, value in data.items():
        if value is not None:
            setattr(questionnaire, key, value)
    
    questionnaire.updated_at = datetime.now()
    session.add(questionnaire)
    session.commit()
    session.refresh(questionnaire)
    return questionnaire


async def delete_questionnaire(session: Session, questionnaire_id: int) -> bool:
    """删除问卷."""
    questionnaire = session.get(Questionnaire, questionnaire_id)
    if not questionnaire:
        return False
    
    session.delete(questionnaire)
    session.commit()
    return True


# ========== 测评管理 ==========

def generate_assessment_code() -> str:
    """生成测评唯一码."""
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    random_str = ''.join(random.choices(string.ascii_uppercase + string.digits, k=4))
    return f"ASSE-{timestamp}-{random_str}"


async def create_assessment(session: Session, data: dict) -> Assessment:
    """创建测评."""
    code = generate_assessment_code()
    assessment_data = {**data, "code": code}
    assessment = Assessment(**assessment_data)
    session.add(assessment)
    session.commit()
    session.refresh(assessment)
    return assessment


async def get_assessments(
    session: Session, skip: int = 0, limit: int = 100
) -> Tuple[List[Assessment], int]:
    """获取测评列表."""
    total = session.scalar(select(func.count()).select_from(Assessment))
    statement = select(Assessment).offset(skip).limit(limit).order_by(Assessment.created_at.desc())
    assessments = session.exec(statement).all()
    return list(assessments), total or 0


async def get_assessment_by_code(session: Session, code: str) -> Optional[Assessment]:
    """根据code获取测评."""
    statement = select(Assessment).where(Assessment.code == code)
    return session.exec(statement).first()


# ========== 提交记录管理 ==========

def generate_submission_code() -> str:
    """生成提交记录唯一码."""
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    random_str = ''.join(random.choices(string.digits, k=3))
    return f"SUB-{timestamp}-{random_str}"


async def check_can_submit(
    session: Session, 
    assessment_id: int, 
    phone: str, 
    name: str = ""
) -> dict:
    """
    检查是否可以提交测评.
    
    返回:
        {
            "can_submit": bool,
            "reason": str,  # 如果不能提交，说明原因
            "submission_number": int,  # 这是第几次提交
            "previous_submissions": list  # 之前的提交记录摘要
        }
    """
    assessment = session.get(Assessment, assessment_id)
    if not assessment:
        return {"can_submit": False, "reason": "测评不存在", "submission_number": 0, "previous_submissions": []}
    
    # 根据 repeat_check_by 确定查询条件
    if assessment.repeat_check_by == "phone_name":
        condition = and_(
            Submission.assessment_id == assessment_id,
            Submission.candidate_phone == phone,
            Submission.candidate_name == name
        )
    else:  # 默认按手机号
        condition = and_(
            Submission.assessment_id == assessment_id,
            Submission.candidate_phone == phone
        )
    
    # 查询该用户在此测评的所有提交记录
    statement = select(Submission).where(condition).order_by(Submission.submitted_at.desc())
    submissions = session.exec(statement).all()
    submission_count = len(submissions)
    
    # 获取之前提交的摘要
    previous_submissions = [
        {
            "code": sub.code,
            "submitted_at": sub.submitted_at.isoformat() if sub.submitted_at else None,
            "status": sub.status,
            "total_score": sub.total_score,
            "grade": sub.grade
        }
        for sub in submissions[:5]  # 只返回最近5条
    ]
    
    # 1. 检查是否允许重复
    if not assessment.allow_repeat and submission_count > 0:
        return {
            "can_submit": False, 
            "reason": "该测评不允许重复提交",
            "submission_number": submission_count,
            "previous_submissions": previous_submissions
        }
    
    # 2. 检查提交间隔
    if assessment.repeat_interval_hours > 0 and submissions:
        last_submission = submissions[0]
        if last_submission.submitted_at:
            hours_since = (datetime.now() - last_submission.submitted_at).total_seconds() / 3600
            if hours_since < assessment.repeat_interval_hours:
                remaining_hours = assessment.repeat_interval_hours - hours_since
                if remaining_hours < 1:
                    remaining_text = f"{int(remaining_hours * 60)}分钟"
                else:
                    remaining_text = f"{int(remaining_hours)}小时"
                return {
                    "can_submit": False, 
                    "reason": f"距上次提交不足{assessment.repeat_interval_hours}小时，请{remaining_text}后再试",
                    "submission_number": submission_count,
                    "previous_submissions": previous_submissions
                }
    
    # 3. 检查提交次数上限
    if assessment.max_submissions > 0 and submission_count >= assessment.max_submissions:
        return {
            "can_submit": False, 
            "reason": f"已达到最大提交次数({assessment.max_submissions}次)",
            "submission_number": submission_count,
            "previous_submissions": previous_submissions
        }
    
    return {
        "can_submit": True,
        "reason": "",
        "submission_number": submission_count + 1,
        "previous_submissions": previous_submissions
    }


async def increment_view_count(session: Session, assessment_id: int) -> None:
    """增加浏览量统计."""
    assessment = session.get(Assessment, assessment_id)
    if assessment:
        assessment.view_count = (assessment.view_count or 0) + 1
        session.add(assessment)
        session.commit()


async def increment_start_count(session: Session, assessment_id: int) -> None:
    """增加开始测评数统计."""
    assessment = session.get(Assessment, assessment_id)
    if assessment:
        assessment.start_count = (assessment.start_count or 0) + 1
        session.add(assessment)
        session.commit()


async def create_submission(session: Session, assessment_id: int, data: dict) -> Submission:
    """创建提交记录（候选人开始测评）."""
    # 获取测评信息
    assessment = session.get(Assessment, assessment_id)
    if not assessment:
        raise ValueError("测评不存在")
    
    code = generate_submission_code()
    
    # ⭐ 提取 custom_data 中的关键字段（如果存在）
    custom_data = data.get("custom_data", {})
    
    # V45: 调试日志 - 查看传入的数据
    print(f"[create_submission] 传入数据: {data}")
    print(f"[create_submission] custom_data: {custom_data}")
    
    # ⭐ 提取应聘岗位（可能在 data、custom_data 或其他字段中）
    # 支持多种字段名：target_position, position, 应聘岗位, text（标签为应聘岗位的自定义字段）
    target_position = (
        data.get("target_position") or 
        data.get("position") or
        custom_data.get("target_position") or
        custom_data.get("position") or
        custom_data.get("text")  # 自定义字段可能用 "text" 作为 name
    )
    # 如果还没找到，遍历 custom_data 找包含"岗位"的值
    if not target_position:
        for key, value in custom_data.items():
            if value and isinstance(value, str) and len(value) < 50:  # 合理长度的岗位名
                # 如果 key 包含 position 或 text，可能是岗位字段
                if 'position' in key.lower() or key.startswith('text'):
                    target_position = value
                    break
    print(f"[create_submission] target_position: {target_position}")
    
    # ⭐ V45: 提取性别（可能在 data 或 custom_data 中）
    gender = data.get("gender") or custom_data.get("gender")
    print(f"[create_submission] gender: {gender}")
    
    # ⭐ 通过手机号+姓名双重校验查找候选人
    candidate_id = None
    candidate_name = data.get("candidate_name", "").strip()
    candidate_phone = data.get("candidate_phone", "").strip()
    
    if candidate_name and candidate_phone:
        # 查找匹配的候选人
        statement = select(Candidate).where(
            and_(
                Candidate.name == candidate_name,
                Candidate.phone == candidate_phone
            )
        )
        candidate = session.exec(statement).first()
        
        if candidate:
            candidate_id = candidate.id
            # 更新候选人的submission_id（关联最新的提交）
            # 这里暂不更新，因为一个候选人可能有多次测评
    
    submission_data = {
        **data,
        "code": code,
        "assessment_id": assessment_id,
        "questionnaire_id": assessment.questionnaire_id,
        "status": "in_progress",
        "candidate_id": candidate_id,  # ⭐ 关联候选人
        "target_position": target_position,  # ⭐ 确保应聘岗位字段正确保存
        "gender": gender,  # ⭐ V45: 确保性别字段正确保存
    }
    
    submission = Submission(**submission_data)
    session.add(submission)
    session.commit()
    session.refresh(submission)
    return submission


async def get_submissions(
    session: Session,
    assessment_id: Optional[int] = None,
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    category: Optional[str] = None
) -> Tuple[List[Submission], int]:
    """获取提交记录列表，支持按问卷category过滤.
    
    Args:
        session: 数据库会话
        assessment_id: 测评ID（可选）
        status: 提交状态（可选）
        skip: 跳过数量
        limit: 返回数量
        category: 问卷分类过滤（professional/scored/survey/custom）
                  'custom' 表示获取所有非professional的问卷（scored + survey）
    """
    # 如果需要按category过滤，先获取符合条件的问卷ID列表
    questionnaire_ids = None
    if category:
        if category == 'custom':
            # ⭐ custom类别：获取所有非professional的问卷（scored + survey）
            q_statement = select(Questionnaire.id).where(Questionnaire.category.in_(['scored', 'survey']))
        else:
            q_statement = select(Questionnaire.id).where(Questionnaire.category == category)
        # ⭐ 修复：select(Questionnaire.id) 返回的是 int，不是对象
        questionnaire_ids = list(session.exec(q_statement).all())
        if not questionnaire_ids:
            return [], 0
    
    statement = select(Submission)
    count_statement = select(func.count()).select_from(Submission)
    
    if assessment_id:
        statement = statement.where(Submission.assessment_id == assessment_id)
        count_statement = count_statement.where(Submission.assessment_id == assessment_id)
    if status:
        statement = statement.where(Submission.status == status)
        count_statement = count_statement.where(Submission.status == status)
    if questionnaire_ids is not None:
        statement = statement.where(Submission.questionnaire_id.in_(questionnaire_ids))
        count_statement = count_statement.where(Submission.questionnaire_id.in_(questionnaire_ids))
    
    total = session.scalar(count_statement)
    
    statement = statement.offset(skip).limit(limit).order_by(Submission.started_at.desc())
    submissions = session.exec(statement).all()
    
    return list(submissions), total or 0


async def get_submission_by_id(session: Session, submission_id: int) -> Optional[Submission]:
    """根据ID获取单个提交记录."""
    statement = select(Submission).where(Submission.id == submission_id)
    return session.exec(statement).first()


async def get_submission_answers(session: Session, submission_id: int) -> dict:
    """获取提交记录的答案数据."""
    from app.models import SubmissionAnswer, Question
    
    # 查询答案记录
    statement = select(SubmissionAnswer).where(SubmissionAnswer.submission_id == submission_id)
    answer_records = session.exec(statement).all()
    
    # 构建答案字典: {question_id: {value, score}}
    answers = {}
    for ans in answer_records:
        # 获取题目信息
        question = session.get(Question, ans.question_id)
        answers[str(ans.question_id)] = {
            "value": ans.value,
            "score": ans.score,
            "question_text": question.text if question else None,
        }
    
    return answers


async def get_candidate_by_submission(session: Session, submission_id: int) -> Optional[dict]:
    """通过提交记录获取候选人信息."""
    from app.models import Candidate
    
    # 尝试通过 submission_id 关联查找候选人
    statement = select(Candidate).where(Candidate.submission_id == submission_id)
    candidate = session.exec(statement).first()
    
    if candidate:
        return {
            "name": candidate.name,
            "phone": candidate.phone,
        }
    
    return None


async def delete_submission(session: Session, submission_id: int) -> bool:
    """删除提交记录."""
    statement = select(Submission).where(Submission.id == submission_id)
    submission = session.exec(statement).first()
    
    if not submission:
        return False
    
    session.delete(submission)
    session.commit()
    return True


async def update_assessment(session: Session, assessment_id: int, data: dict) -> Optional[Assessment]:
    """更新测评配置."""
    assessment = session.get(Assessment, assessment_id)
    
    if not assessment:
        return None
    
    # 更新字段
    for key, value in data.items():
        if value is not None and hasattr(assessment, key):
            setattr(assessment, key, value)
    
    assessment.updated_at = datetime.now()
    session.add(assessment)
    session.commit()
    session.refresh(assessment)
    
    return assessment


async def delete_assessment(
    session: Session, 
    assessment_id: int,
    force_delete_submissions: bool = False
) -> dict:
    """
    删除测评（分发链接）.
    
    Args:
        session: 数据库会话
        assessment_id: 测评ID
        force_delete_submissions: 是否强制删除关联的提交记录
            - False: 如果有提交记录，返回错误信息，不删除
            - True: 删除分发链接及所有关联的提交记录
    
    Returns:
        dict: 包含删除结果的字典
    """
    assessment = session.get(Assessment, assessment_id)
    
    if not assessment:
        return {"success": False, "error": "测评不存在"}
    
    # 检查是否有关联的提交记录
    statement = select(Submission).where(Submission.assessment_id == assessment_id)
    submissions = session.exec(statement).all()
    submission_count = len(submissions)
    
    if submission_count > 0 and not force_delete_submissions:
        # 有提交记录但未强制删除，返回警告
        return {
            "success": False,
            "error": "has_submissions",
            "submission_count": submission_count,
            "message": f"该分发链接下有 {submission_count} 条提交记录，删除后数据将无法恢复。请确认是否继续删除？"
        }
    
    # 执行删除
    deleted_submissions = 0
    if submission_count > 0:
        for sub in submissions:
            session.delete(sub)
            deleted_submissions = submission_count
    
    session.delete(assessment)
    session.commit()
    
    return {
        "success": True,
        "deleted_submissions": deleted_submissions,
        "message": f"删除成功" + (f"，同时删除了 {deleted_submissions} 条提交记录" if deleted_submissions > 0 else "")
    }


async def submit_answers(session: Session, submission_code: str, answers: dict) -> Submission:
    """提交答案并计算得分."""
    statement = select(Submission).where(Submission.code == submission_code)
    submission = session.exec(statement).first()
    
    if not submission:
        raise ValueError("提交记录不存在")
    
    if submission.status == "completed":
        raise ValueError("该测评已完成，无法重复提交")
    
    # 获取问卷信息
    questionnaire = session.get(Questionnaire, submission.questionnaire_id)
    if not questionnaire:
        raise ValueError("问卷不存在")
    
    # ⭐ 根据问卷类型调用对应的评分算法
    try:
        questionnaire_type = questionnaire.type.upper() if questionnaire.type else ''
        
        if questionnaire_type in ['MBTI', 'DISC', 'EPQ']:
            # 专业测评：使用专业评分算法
            # 获取题目数据以便评分算法能根据维度评分
            questions = questionnaire.questions_data.get('questions', [])
            result = score_professional_assessment(questionnaire_type, answers, questions)
            
            # 构建result_details用于前端展示
            if questionnaire_type == 'MBTI':
                result_details = {
                    'mbti_type': result.get('mbti_type'),
                    'mbti_description': result.get('mbti_description'),
                    'mbti_dimensions': result.get('mbti_dimensions')
                }
            elif questionnaire_type == 'DISC':
                result_details = {
                    'disc_type': result.get('disc_type'),
                    'disc_description': result.get('disc_description'),
                    'disc_dimensions': result.get('disc_dimensions')
                }
            elif questionnaire_type == 'EPQ':
                result_details = {
                    'personality_trait': result.get('personality_trait'),
                    'dimensions': result.get('dimensions')
                }
            
            submission.result_details = result_details
            submission.scores = result.get('raw_scores') or result.get('dimensions', {})
            submission.total_score = result.get('total_score', 0)
            submission.grade = result.get('grade', 'C')
            
        else:
            # 自定义问卷：使用新的评分算法
            questionnaire_dict = {
                "custom_type": questionnaire.custom_type,
                "scoring_config": questionnaire.scoring_config,
                "questions_data": questionnaire.questions_data
            }
            
            # 转换答案格式
            answers_list = []
            if isinstance(answers, dict):
                for q_id, answer_data in answers.items():
                    answers_list.append({
                        "question_id": q_id,
                        "answer": answer_data
                    })
            elif isinstance(answers, list):
                answers_list = answers
            
            # 使用新的评分算法
            result = calculate_custom_questionnaire_score(questionnaire_dict, answers_list)
            
            # 保存结果
            submission.result_details = {
                "custom_type": questionnaire.custom_type,
                "answers": result.get("detailed_answers", [])
            }
            submission.total_score = result.get("total_score")
            submission.max_score = result.get("max_score")
            submission.score_percentage = result.get("score_percentage")
            submission.grade = result.get("grade")
            submission.scores = {}  # 详细得分已在result_details中
        
        # 保存答案和状态
        submission.answers = answers
        submission.status = "completed"
        submission.submitted_at = datetime.now()
        
        # ⭐ 创建或关联候选人记录
        candidate = await _get_or_create_candidate(
            session, 
            submission.candidate_name,
            submission.candidate_phone,
            submission.candidate_email,
            submission.target_position,
            submission.gender  # V45: 传递性别
        )
        if candidate:
            submission.candidate_id = candidate.id
        
        session.add(submission)
        session.commit()
        session.refresh(submission)
        
    except ProfessionalScoringError as e:
        raise ValueError(f"评分失败: {str(e)}")
    except Exception as e:
        raise ValueError(f"提交失败: {str(e)}")
    
    return submission


async def _get_or_create_candidate(
    session: Session,
    name: str,
    phone: str,
    email: Optional[str] = None,
    position: Optional[str] = None,
    gender: Optional[str] = None  # V45: 添加性别参数
) -> Optional[Candidate]:
    """根据手机号获取或创建候选人记录.
    
    逻辑：
    1. 首先根据手机号查找已存在的候选人
    2. 如果存在，更新其信息（姓名、邮箱、岗位、性别）
    3. 如果不存在，创建新的候选人记录
    """
    if not phone:
        return None
    
    try:
        # 查找已存在的候选人
        statement = select(Candidate).where(Candidate.phone == phone)
        existing_candidate = session.exec(statement).first()
        
        if existing_candidate:
            # 更新候选人信息
            if name and name != existing_candidate.name:
                existing_candidate.name = name
            if email and email != existing_candidate.email:
                existing_candidate.email = email
            if position and position != existing_candidate.position:
                existing_candidate.position = position
            # V45: 更新性别
            if gender and gender != getattr(existing_candidate, 'gender', None):
                existing_candidate.gender = gender
            existing_candidate.updated_at = datetime.now()
            session.add(existing_candidate)
            return existing_candidate
        else:
            # 创建新候选人
            new_candidate = Candidate(
                name=name or "未知",
                phone=phone,
                email=email,
                position=position,
                gender=gender,  # V45: 保存性别
                status="completed"  # 已完成测评
            )
            session.add(new_candidate)
            session.flush()  # 获取ID但不提交
            return new_candidate
            
    except Exception as e:
        print(f"[候选人关联] 创建/更新候选人失败: {e}")
        return None


# ========== 统计相关 ==========

async def get_submission_statistics(
    session: Session,
    category: Optional[str] = None,
    questionnaire_id: Optional[int] = None
) -> dict:
    """获取提交记录统计数据."""
    from sqlmodel import select, func
    
    # 构建查询基础
    base_query = select(Submission).where(Submission.status == "completed")
    
    # 如果指定了问卷ID
    if questionnaire_id:
        base_query = base_query.where(Submission.questionnaire_id == questionnaire_id)
    
    # 如果指定了category，需要先获取对应的问卷ID列表
    if category:
        q_statement = select(Questionnaire.id).where(Questionnaire.category == category)
        questionnaire_ids = session.exec(q_statement).all()
        if questionnaire_ids:
            base_query = base_query.where(Submission.questionnaire_id.in_(questionnaire_ids))
        else:
            return {
                "total_submissions": 0,
                "average_score": 0,
                "pass_rate": 0,
                "grade_distribution": {"A": 0, "B": 0, "C": 0, "D": 0},
                "submissions": []
            }
    
    # 执行查询
    all_submissions = session.exec(base_query).all()
    
    # 计算统计数据
    total = len(all_submissions)
    
    if total == 0:
        return {
            "total_submissions": 0,
            "average_score": 0,
            "pass_rate": 0,
            "grade_distribution": {"A": 0, "B": 0, "C": 0, "D": 0},
            "submissions": []
        }
    
    # 计算平均分（过滤掉 None 值）
    valid_scores = [s.total_score for s in all_submissions if s.total_score is not None]
    average_score = sum(valid_scores) / len(valid_scores) if valid_scores else 0
    
    # 计算合格率（假设60分及格）
    pass_count = len([s for s in valid_scores if s >= 60])
    pass_rate = (pass_count / len(valid_scores) * 100) if valid_scores else 0
    
    # 计算等级分布
    grade_distribution = {"A": 0, "B": 0, "C": 0, "D": 0}
    for s in all_submissions:
        grade = (s.grade or "D").upper()
        if grade in grade_distribution:
            grade_distribution[grade] += 1
    
    # 构建返回数据
    return {
        "total_submissions": total,
        "average_score": round(average_score, 2),
        "pass_rate": round(pass_rate, 2),
        "grade_distribution": grade_distribution,
        "grade_percentages": {
            grade: round(count / total * 100, 1) if total > 0 else 0
            for grade, count in grade_distribution.items()
        },
        "submissions": [
            {
                "id": s.id,
                "candidate_name": s.candidate_name,
                "candidate_phone": s.candidate_phone,
                "total_score": s.total_score,
                "grade": s.grade,
                "submitted_at": s.submitted_at.isoformat() if s.submitted_at else None
            }
            for s in all_submissions[:100]  # 限制返回数量
        ]
    }


def _get_scale_label(score: int, scale_min: int, scale_max: int, min_label: str, max_label: str) -> str:
    """
    V46: 智能生成量表题的描述文本
    
    支持多种常见量表类型：满意度、同意度、频率、程度等
    """
    total_levels = scale_max - scale_min + 1
    position = score - scale_min  # 0-based position
    
    # 预定义的量表描述模板
    SCALE_TEMPLATES = {
        ('满意', 5): ['非常不满意', '不太满意', '一般', '比较满意', '非常满意'],
        ('满意', 4): ['不满意', '一般', '满意', '非常满意'],
        ('同意', 5): ['非常不同意', '不同意', '一般', '同意', '非常同意'],
        ('频率', 5): ['从不', '很少', '有时', '经常', '总是'],
        ('符合', 5): ['完全不符合', '不太符合', '一般', '比较符合', '完全符合'],
        ('重要', 5): ['非常不重要', '不太重要', '一般', '比较重要', '非常重要'],
    }
    
    labels = None
    for (keyword, levels), template in SCALE_TEMPLATES.items():
        if levels == total_levels and (keyword in min_label or keyword in max_label):
            labels = template
            break
    
    if not labels and min_label and max_label:
        if total_levels == 5:
            labels = [min_label, f'偏向{min_label[:2]}', '一般', f'偏向{max_label[:2]}', max_label]
        elif total_levels == 3:
            labels = [min_label, '一般', max_label]
        else:
            if score == scale_min:
                return f"{score}分 ({min_label})"
            elif score == scale_max:
                return f"{score}分 ({max_label})"
            return f"{score}分"
    
    if not labels:
        if total_levels == 5:
            labels = ['很低', '较低', '一般', '较高', '很高']
        else:
            return f"{score}分"
    
    if labels and 0 <= position < len(labels):
        return f"{score}分 ({labels[position]})"
    return f"{score}分"


async def get_question_answer_statistics(
    session: Session,
    questionnaire_id: int
) -> dict:
    """
    V42: 获取问卷的题目答案统计数据.
    
    返回每道题的选项分布统计，用于问卷统计页面的数据可视化。
    """
    from sqlmodel import select
    from collections import Counter, defaultdict
    from datetime import datetime, timedelta
    
    # 获取问卷信息
    questionnaire = session.get(Questionnaire, questionnaire_id)
    if not questionnaire:
        return {"error": "问卷不存在", "questions": []}
    
    # 获取所有提交记录（包括已完成和进行中的）
    # V46: 修复统计数据不显示问题 - 同时查询多种完成状态
    query = select(Submission).where(
        Submission.questionnaire_id == questionnaire_id,
        Submission.status.in_(["completed", "已完成", "done", "submitted"])
    ).order_by(Submission.submitted_at.desc())
    
    submissions = list(session.exec(query).all())
    
    # 如果没找到，尝试不限制状态查询
    if not submissions:
        query_all = select(Submission).where(
            Submission.questionnaire_id == questionnaire_id
        ).order_by(Submission.submitted_at.desc())
        all_submissions = list(session.exec(query_all).all())
        # 过滤出有答案的提交
        submissions = [s for s in all_submissions if s.answers and len(s.answers) > 0]
    
    total_submissions = len(submissions)
    
    if total_submissions == 0:
        return {
            "questionnaire_id": questionnaire_id,
            "questionnaire_name": questionnaire.name,
            "questionnaire_type": questionnaire.type,
            "questionnaire_category": questionnaire.category,
            "total_submissions": 0,
            "completion_rate": 0,
            "average_score": None,
            "average_duration_minutes": None,
            "questions": [],
            "daily_trend": [],
            "grade_distribution": {"A": 0, "B": 0, "C": 0, "D": 0}
        }
    
    # 解析问卷题目
    questions_data = questionnaire.questions_data or []
    if isinstance(questions_data, str):
        import json
        try:
            questions_data = json.loads(questions_data)
        except:
            questions_data = []
    
    # 处理 {"questions": [...]} 格式
    if isinstance(questions_data, dict):
        questions_data = questions_data.get("questions", [])
    
    # 确保是列表
    if not isinstance(questions_data, list):
        questions_data = []
    
    # 统计每道题的答案分布
    question_stats = []
    
    for q_idx, question in enumerate(questions_data):
        # 跳过非字典类型的项
        if not isinstance(question, dict):
            continue
        
        q_id = question.get("id", str(q_idx + 1))
        q_text = question.get("text", question.get("question", f"问题 {q_idx + 1}"))
        q_type = question.get("type", "single")  # single, multiple, text, rating
        options = question.get("options", [])
        
        # 收集所有答案
        answer_counts = Counter()
        text_answers = []
        
        for sub in submissions:
            answers = sub.answers or {}
            if isinstance(answers, str):
                import json
                try:
                    answers = json.loads(answers)
                except:
                    answers = {}
            
            answer = answers.get(q_id) or answers.get(str(q_idx))
            
            if answer is None:
                continue
            
            if q_type in ("text", "textarea"):
                # 文本题收集答案
                if isinstance(answer, str) and answer.strip():
                    text_answers.append(answer.strip())
            elif q_type in ("multiple", "checkbox") or isinstance(answer, list):
                # 多选题
                for a in (answer if isinstance(answer, list) else [answer]):
                    answer_counts[str(a)] += 1
            else:
                # 单选题
                answer_counts[str(answer)] += 1
        
        # 构建选项统计
        option_stats = []
        total_answers = sum(answer_counts.values()) if answer_counts else len(text_answers)
        
        if q_type == "text" or q_type == "textarea":
            # 文本题：显示部分回答样本
            option_stats = [
                {"text": ans, "count": 1}
                for ans in text_answers[:10]  # 只显示前10个
            ]
        elif q_type == "scale" or q_type == "rating":
            # V46: 量表题/评分题 - 根据 scale 配置生成选项统计
            scale_config = question.get("scale", {})
            scale_min = scale_config.get("min", 1)
            scale_max = scale_config.get("max", 5)
            min_label = scale_config.get("minLabel", "")
            max_label = scale_config.get("maxLabel", "")
            
            for score in range(scale_min, scale_max + 1):
                count = answer_counts.get(str(score), 0)
                percentage = round(count / total_submissions * 100, 1) if total_submissions > 0 else 0
                
                # 使用智能标签生成
                label = _get_scale_label(score, scale_min, scale_max, min_label, max_label)
                
                option_stats.append({
                    "index": score - scale_min,
                    "text": label,
                    "count": count,
                    "percentage": percentage
                })
        else:
            # 选择题：统计每个选项的选择次数
            for opt_idx, opt in enumerate(options):
                if isinstance(opt, dict):
                    opt_text = opt.get("text", opt.get("label", str(opt_idx)))
                    opt_value = str(opt.get("value", opt_idx))
                else:
                    opt_text = str(opt)
                    opt_value = str(opt_idx)
                
                # 确保 opt_text 是字符串
                if not isinstance(opt_text, str):
                    opt_text = str(opt_text)
                
                # 尝试匹配答案（可能是索引、值或文本）
                count = answer_counts.get(str(opt_idx), 0)
                count += answer_counts.get(opt_value, 0)
                if opt_text != opt_value and opt_text != str(opt_idx):
                    count += answer_counts.get(opt_text, 0)
                
                percentage = round(count / total_submissions * 100, 1) if total_submissions > 0 else 0
                
                option_stats.append({
                    "index": opt_idx,
                    "text": opt_text,
                    "count": count,
                    "percentage": percentage
                })
        
        question_stats.append({
            "id": q_id,
            "index": q_idx + 1,
            "text": q_text,
            "type": q_type,
            "total_answers": total_answers,
            "options": option_stats
        })
    
    # 计算平均分（仅评分问卷）
    average_score = None
    if questionnaire.category == "scored":
        valid_scores = [s.total_score for s in submissions if s.total_score is not None]
        if valid_scores:
            average_score = round(sum(valid_scores) / len(valid_scores), 1)
    
    # 计算平均用时
    average_duration = None
    durations = []
    for sub in submissions:
        if sub.started_at and sub.submitted_at:
            duration = (sub.submitted_at - sub.started_at).total_seconds() / 60
            if 0 < duration < 120:  # 排除异常值
                durations.append(duration)
    if durations:
        average_duration = round(sum(durations) / len(durations), 1)
    
    # 计算每日提交趋势（最近7天）
    daily_trend = []
    today = datetime.now().date()
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        count = sum(1 for s in submissions if s.submitted_at and s.submitted_at.date() == day)
        daily_trend.append({
            "date": day.strftime("%m/%d"),
            "count": count
        })
    
    # 等级分布
    grade_distribution = {"A": 0, "B": 0, "C": 0, "D": 0}
    for sub in submissions:
        grade = (sub.grade or "D").upper()
        if grade in grade_distribution:
            grade_distribution[grade] += 1
    
    return {
        "questionnaire_id": questionnaire_id,
        "questionnaire_name": questionnaire.name,
        "questionnaire_type": questionnaire.type,
        "questionnaire_category": questionnaire.category,
        "total_submissions": total_submissions,
        "completion_rate": 100,  # 只统计已完成的
        "average_score": average_score,
        "average_duration_minutes": average_duration,
        "questions": question_stats,
        "daily_trend": daily_trend,
        "grade_distribution": grade_distribution,
        "grade_percentages": {
            grade: round(count / total_submissions * 100, 1) if total_submissions > 0 else 0
            for grade, count in grade_distribution.items()
        }
    }


async def export_submissions_to_excel(
    session: Session,
    category: Optional[str] = None,
    questionnaire_id: Optional[int] = None
) -> bytes:
    """导出提交记录为Excel文件."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from sqlmodel import select
    
    # 构建查询
    query = select(Submission)
    
    if questionnaire_id:
        query = query.where(Submission.questionnaire_id == questionnaire_id)
    
    if category:
        q_statement = select(Questionnaire.id).where(Questionnaire.category == category)
        questionnaire_ids = session.exec(q_statement).all()
        if questionnaire_ids:
            query = query.where(Submission.questionnaire_id.in_(questionnaire_ids))
    
    submissions = session.exec(query).all()
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "提交记录"
    
    # 定义样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # 写入表头
    headers = ["序号", "姓名", "电话", "问卷", "得分", "等级", "状态", "提交时间"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 设置列宽
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 20
    
    # 写入数据
    for row, sub in enumerate(submissions, 2):
        ws.cell(row=row, column=1, value=row - 1).border = thin_border
        ws.cell(row=row, column=2, value=sub.candidate_name or "").border = thin_border
        ws.cell(row=row, column=3, value=sub.candidate_phone or "").border = thin_border
        ws.cell(row=row, column=4, value=sub.questionnaire_name or "").border = thin_border
        ws.cell(row=row, column=5, value=sub.total_score).border = thin_border
        ws.cell(row=row, column=6, value=sub.grade or "").border = thin_border
        ws.cell(row=row, column=7, value="已完成" if sub.status == "completed" else "进行中").border = thin_border
        ws.cell(row=row, column=8, value=sub.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if sub.submitted_at else "").border = thin_border
    
    # 保存到字节流
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

