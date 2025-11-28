"""
TalentLens 后端服务
智能人才洞察平台 - Flask API
"""
from flask import Flask, request, jsonify, session
from flask_cors import CORS
import os
import json
import requests
from datetime import datetime
from functools import wraps
from config import Config

app = Flask(__name__)
app.secret_key = os.urandom(24)
app.config.from_object(Config)
CORS(app, supports_credentials=True)

# 初始化配置
Config.init_app(app)

# ==================== 认证装饰器 ====================
def login_required(f):
    """登录验证装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return jsonify({'success': False, 'message': '未登录'}), 401
        return f(*args, **kwargs)
    return decorated_function

# ==================== 认证接口 ====================
@app.route('/api/auth/login', methods=['POST'])
def login():
    """管理员登录"""
    data = request.get_json()
    password = data.get('password', '')
    
    if password == Config.ADMIN_PASSWORD:
        session['logged_in'] = True
        session['login_time'] = datetime.now().isoformat()
        return jsonify({
            'success': True,
            'message': '登录成功'
        })
    
    return jsonify({
        'success': False,
        'message': '密码错误'
    }), 401

@app.route('/api/auth/logout', methods=['POST'])
def logout():
    """退出登录"""
    session.clear()
    return jsonify({'success': True, 'message': '已退出'})

@app.route('/api/auth/status', methods=['GET'])
def auth_status():
    """检查登录状态"""
    return jsonify({
        'logged_in': session.get('logged_in', False),
        'login_time': session.get('login_time')
    })

# ==================== AI 分析接口 ====================
@app.route('/api/ai/analyze', methods=['POST'])
@login_required
def ai_analyze():
    """AI 智能分析候选人"""
    data = request.get_json()
    candidate = data.get('candidate', {})
    analysis_type = data.get('type', 'personality')
    
    if not Config.DEEPSEEK_API_KEY:
        return jsonify({
            'success': False,
            'message': 'AI 服务未配置，请设置 DEEPSEEK_API_KEY'
        }), 500
    
    # 构建提示词
    prompt = build_analysis_prompt(candidate, analysis_type)
    
    try:
        result = call_deepseek_api(prompt)
        return jsonify({
            'success': True,
            'analysis': result
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'AI 分析失败: {str(e)}'
        }), 500

def build_analysis_prompt(candidate, analysis_type):
    """构建分析提示词"""
    name = candidate.get('name', '候选人')
    position = candidate.get('position', '未知岗位')
    scores = candidate.get('scores', {})
    
    e_score = scores.get('E', 0)
    n_score = scores.get('N', 0)
    p_score = scores.get('P', 0)
    l_score = scores.get('L', 0)
    
    base_info = f"""
候选人信息：
- 姓名：{name}
- 应聘岗位：{position}
- EPQ测评结果：
  - 外向性(E)：{e_score}/24
  - 神经质(N)：{n_score}/24
  - 精神质(P)：{p_score}/24
  - 掩饰性(L)：{l_score}/24
"""
    
    if analysis_type == 'personality':
        return f"""你是一位资深的人力资源专家和心理学家。请根据以下候选人的EPQ测评结果，提供专业的性格分析报告。

{base_info}

请从以下几个方面进行分析（使用中文回复，控制在300字以内）：
1. 性格特质总结（一句话概括）
2. 核心优势（2-3点）
3. 潜在风险（1-2点）
4. 工作风格特点
5. 与{position}岗位的匹配度评估

请使用专业但易懂的语言，避免过于学术化。"""

    elif analysis_type == 'interview':
        return f"""你是一位资深的面试官和人力资源专家。请根据以下候选人的EPQ测评结果，提供针对性的面试问题建议。

{base_info}

请提供5个针对性的面试问题，每个问题需要：
1. 问题本身
2. 设计意图（为什么问这个问题）
3. 理想回答方向

问题应该能够验证测评结果的准确性，并深入了解候选人的真实特质。"""

    elif analysis_type == 'development':
        return f"""你是一位职业发展顾问和人才发展专家。请根据以下候选人的EPQ测评结果，提供职业发展建议。

{base_info}

请从以下方面提供建议（使用中文回复，控制在250字以内）：
1. 适合的职业发展方向
2. 需要提升的能力
3. 职业发展建议
4. 团队协作建议"""

    else:
        return f"""请分析以下候选人的信息：{base_info}"""

def call_deepseek_api(prompt):
    """调用 DeepSeek API"""
    headers = {
        'Authorization': f'Bearer {Config.DEEPSEEK_API_KEY}',
        'Content-Type': 'application/json'
    }
    
    payload = {
        'model': 'deepseek-chat',
        'messages': [
            {'role': 'system', 'content': '你是一位专业的人力资源顾问和心理学专家，擅长人才评估和职业发展建议。'},
            {'role': 'user', 'content': prompt}
        ],
        'temperature': 0.7,
        'max_tokens': 1000
    }
    
    response = requests.post(
        Config.DEEPSEEK_API_URL,
        headers=headers,
        json=payload,
        timeout=30
    )
    
    if response.status_code != 200:
        raise Exception(f'API 请求失败: {response.status_code}')
    
    result = response.json()
    return result['choices'][0]['message']['content']

# ==================== 简历解析接口 ====================
@app.route('/api/resume/upload', methods=['POST'])
@login_required
def upload_resume():
    """上传并解析简历"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未找到文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择文件'}), 400
    
    # 检查文件类型
    ext = file.filename.rsplit('.', 1)[-1].lower() if '.' in file.filename else ''
    if ext not in Config.ALLOWED_EXTENSIONS:
        return jsonify({'success': False, 'message': f'不支持的文件类型: {ext}'}), 400
    
    # 保存文件
    filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    filepath = os.path.join(Config.UPLOAD_FOLDER, filename)
    file.save(filepath)
    
    # 解析简历
    try:
        parsed_data = parse_resume(filepath, ext)
        return jsonify({
            'success': True,
            'filename': filename,
            'parsed': parsed_data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'解析失败: {str(e)}'
        }), 500

def parse_resume(filepath, ext):
    """解析简历内容"""
    text = ''
    
    if ext == 'pdf':
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(filepath)
            for page in reader.pages:
                text += page.extract_text() or ''
        except Exception as e:
            text = f'[PDF解析失败: {str(e)}]'
    
    elif ext in ['png', 'jpg', 'jpeg']:
        try:
            import pytesseract
            from PIL import Image
            image = Image.open(filepath)
            text = pytesseract.image_to_string(image, lang='chi_sim+eng')
        except Exception as e:
            text = f'[图片OCR失败: {str(e)}]'
    
    elif ext in ['xlsx', 'xls']:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    text += ' '.join(str(cell) for cell in row if cell) + '\n'
        except Exception as e:
            text = f'[Excel解析失败: {str(e)}]'
    
    elif ext in ['doc', 'docx']:
        text = '[Word文档解析需要额外依赖，暂不支持]'
    
    # 使用 AI 提取结构化信息
    if text and Config.DEEPSEEK_API_KEY:
        try:
            structured = extract_resume_info(text)
            return {
                'raw_text': text[:2000],  # 限制长度
                'structured': structured
            }
        except:
            pass
    
    return {
        'raw_text': text[:2000],
        'structured': None
    }

def extract_resume_info(text):
    """使用 AI 提取简历结构化信息"""
    prompt = f"""请从以下简历文本中提取结构化信息，以JSON格式返回：

{text[:3000]}

请提取以下字段（如果找不到则返回null）：
{{
    "name": "姓名",
    "phone": "手机号",
    "email": "邮箱",
    "education": "最高学历",
    "school": "毕业院校",
    "major": "专业",
    "work_years": "工作年限",
    "skills": ["技能1", "技能2"],
    "experience_summary": "工作经历摘要（100字以内）"
}}

只返回JSON，不要其他内容。"""
    
    result = call_deepseek_api(prompt)
    
    # 尝试解析JSON
    try:
        # 清理可能的markdown标记
        result = result.strip()
        if result.startswith('```'):
            result = result.split('\n', 1)[1]
        if result.endswith('```'):
            result = result.rsplit('```', 1)[0]
        return json.loads(result)
    except:
        return None

# ==================== 健康检查 ====================
@app.route('/api/health', methods=['GET'])
def health_check():
    """健康检查"""
    return jsonify({
        'status': 'ok',
        'version': '1.0.0',
        'ai_configured': bool(Config.DEEPSEEK_API_KEY)
    })

# ==================== 启动服务 ====================
if __name__ == '__main__':
    print(f"""
╔═══════════════════════════════════════════════════════════╗
║                                                           ║
║   🎯 TalentLens 后端服务启动                              ║
║                                                           ║
║   地址: http://{Config.HOST}:{Config.PORT}                           ║
║   AI状态: {'✅ 已配置' if Config.DEEPSEEK_API_KEY else '❌ 未配置 (设置 DEEPSEEK_API_KEY)'}                  ║
║                                                           ║
╚═══════════════════════════════════════════════════════════╝
""")
    app.run(
        host=Config.HOST,
        port=Config.PORT,
        debug=Config.DEBUG
    )

